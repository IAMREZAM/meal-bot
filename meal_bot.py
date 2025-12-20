import os
import logging
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters
)
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json

# تنظیمات لاگ
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# آیدی ادمین - اینجا را تغییر دهید
ADMIN_ID = 166152961  # آیدی عددی تلگرام ادمین

# States برای ConversationHandler
(ADD_MEAL, ADD_DESSERT, ADD_USER_ID, ADD_USER_NAME, 
 SELECT_DAY_MEAL, SELECT_DAY_DESSERT, BROADCAST_MSG) = range(7)

# Database initialization
def init_db():
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    # جدول کاربران
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (user_id INTEGER PRIMARY KEY,
                  first_name TEXT NOT NULL,
                  last_name TEXT NOT NULL,
                  is_active INTEGER DEFAULT 1,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    # جدول غذاها
    c.execute('''CREATE TABLE IF NOT EXISTS meals
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  type TEXT NOT NULL,
                  day_of_week INTEGER NOT NULL,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    # جدول رزروها
    c.execute('''CREATE TABLE IF NOT EXISTS reservations
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id INTEGER NOT NULL,
                  meal_id INTEGER,
                  dessert_id INTEGER,
                  reservation_date DATE NOT NULL,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  FOREIGN KEY (user_id) REFERENCES users (user_id),
                  FOREIGN KEY (meal_id) REFERENCES meals (id),
                  FOREIGN KEY (dessert_id) REFERENCES meals (id),
                  UNIQUE(user_id, reservation_date))''')
    
    # جدول تنظیمات
    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (key TEXT PRIMARY KEY,
                  value TEXT NOT NULL)''')
    
    conn.commit()
    conn.close()

# تابع کمکی برای بررسی ادمین بودن
def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID

# تابع کمکی برای بررسی کاربر مجاز
def is_authorized_user(user_id: int) -> bool:
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    c.execute("SELECT is_active FROM users WHERE user_id = ?", (user_id,))
    result = c.fetchone()
    conn.close()
    return result is not None and result[0] == 1

# دستور start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    
    if is_admin(user.id):
        keyboard = [
            [InlineKeyboardButton("👥 مدیریت کاربران", callback_data='admin_users')],
            [InlineKeyboardButton("🍽 مدیریت غذاها", callback_data='admin_meals')],
            [InlineKeyboardButton("📊 مشاهده رزروها", callback_data='admin_view_reservations')],
            [InlineKeyboardButton("📥 دریافت فایل اکسل", callback_data='admin_export_excel')],
            [InlineKeyboardButton("📢 ارسال پیام همگانی", callback_data='admin_broadcast')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            f"سلام ادمین عزیز {user.first_name}!\n\n"
            "به پنل مدیریت ربات رزرو غذا خوش آمدید.",
            reply_markup=reply_markup
        )
    elif is_authorized_user(user.id):
        keyboard = [
            [InlineKeyboardButton("🍽 رزرو غذا", callback_data='reserve_food')],
            [InlineKeyboardButton("📋 رزروهای من", callback_data='my_reservations')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            f"سلام {user.first_name}!\n\n"
            "به ربات رزرو غذا خوش آمدید.",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            "متأسفانه شما دسترسی به این ربات ندارید.\n"
            "لطفاً با مدیر سیستم تماس بگیرید."
        )

# منوی اصلی ادمین
async def admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("👥 مدیریت کاربران", callback_data='admin_users')],
        [InlineKeyboardButton("🍽 مدیریت غذاها", callback_data='admin_meals')],
        [InlineKeyboardButton("📊 مشاهده رزروها", callback_data='admin_view_reservations')],
        [InlineKeyboardButton("📥 دریافت فایل اکسل", callback_data='admin_export_excel')],
        [InlineKeyboardButton("📢 ارسال پیام همگانی", callback_data='admin_broadcast')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "پنل مدیریت ربات رزرو غذا:",
        reply_markup=reply_markup
    )

# مدیریت کاربران
async def admin_users_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("➕ افزودن کاربر", callback_data='add_user')],
        [InlineKeyboardButton("📋 لیست کاربران", callback_data='list_users')],
        [InlineKeyboardButton("🔙 بازگشت", callback_data='back_to_admin')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "مدیریت کاربران:",
        reply_markup=reply_markup
    )

# شروع افزودن کاربر
async def start_add_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "لطفاً آیدی عددی کاربر را وارد کنید:\n\n"
        "برای لغو /cancel را ارسال کنید."
    )
    
    return ADD_USER_ID

# دریافت آیدی کاربر
async def receive_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = int(update.message.text)
        context.user_data['new_user_id'] = user_id
        
        await update.message.reply_text(
            "حالا نام و نام خانوادگی کاربر را وارد کنید:\n"
            "مثال: علی احمدی"
        )
        
        return ADD_USER_NAME
    except ValueError:
        await update.message.reply_text(
            "آیدی باید عدد باشد. دوباره تلاش کنید:"
        )
        return ADD_USER_ID

# دریافت نام کاربر
async def receive_user_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name_parts = update.message.text.strip().split(' ', 1)
    
    if len(name_parts) < 2:
        await update.message.reply_text(
            "لطفاً نام و نام خانوادگی را با فاصله وارد کنید:\n"
            "مثال: علی احمدی"
        )
        return ADD_USER_NAME
    
    first_name, last_name = name_parts[0], name_parts[1]
    user_id = context.user_data['new_user_id']
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    try:
        c.execute(
            "INSERT INTO users (user_id, first_name, last_name) VALUES (?, ?, ?)",
            (user_id, first_name, last_name)
        )
        conn.commit()
        await update.message.reply_text(
            f"✅ کاربر {first_name} {last_name} با موفقیت اضافه شد!\n\n"
            "برای بازگشت به منو /start را ارسال کنید."
        )
    except sqlite3.IntegrityError:
        await update.message.reply_text(
            "❌ این کاربر قبلاً ثبت شده است.\n\n"
            "برای بازگشت به منو /start را ارسال کنید."
        )
    finally:
        conn.close()
    
    return ConversationHandler.END

# لیست کاربران
async def list_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    c.execute("SELECT user_id, first_name, last_name, is_active FROM users ORDER BY first_name")
    users = c.fetchall()
    conn.close()
    
    if not users:
        text = "هیچ کاربری ثبت نشده است."
    else:
        text = "📋 لیست کاربران:\n\n"
        for user_id, first_name, last_name, is_active in users:
            status = "✅ فعال" if is_active else "❌ غیرفعال"
            text += f"• {first_name} {last_name} ({user_id}) - {status}\n"
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data='admin_users')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)

# مدیریت غذاها
async def admin_meals_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("➕ افزودن غذا", callback_data='add_meal')],
        [InlineKeyboardButton("➕ افزودن دسر", callback_data='add_dessert')],
        [InlineKeyboardButton("📋 لیست غذاها", callback_data='list_meals')],
        [InlineKeyboardButton("🔙 بازگشت", callback_data='back_to_admin')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "مدیریت غذاها و دسرها:",
        reply_markup=reply_markup
    )

# انتخاب روز برای غذا
async def select_day_for_meal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    context.user_data['meal_type'] = 'meal'
    
    days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    keyboard = []
    for i, day in enumerate(days):
        keyboard.append([InlineKeyboardButton(day, callback_data=f'day_meal_{i}')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin_meals')])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        "روز هفته را انتخاب کنید:",
        reply_markup=reply_markup
    )

# انتخاب روز برای دسر
async def select_day_for_dessert(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    context.user_data['meal_type'] = 'dessert'
    
    days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    keyboard = []
    for i, day in enumerate(days):
        keyboard.append([InlineKeyboardButton(day, callback_data=f'day_dessert_{i}')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin_meals')])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        "روز هفته را انتخاب کنید:",
        reply_markup=reply_markup
    )

# دریافت نام غذا
async def receive_meal_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    day = int(query.data.split('_')[-1])
    meal_type = context.user_data.get('meal_type', 'meal')
    
    context.user_data['meal_day'] = day
    
    days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    meal_type_fa = 'غذا' if meal_type == 'meal' else 'دسر'
    
    await query.edit_message_text(
        f"نام {meal_type_fa} برای روز {days[day]} را وارد کنید:\n\n"
        "برای لغو /cancel را ارسال کنید."
    )
    
    return ADD_MEAL if meal_type == 'meal' else ADD_DESSERT

# ذخیره غذا یا دسر
async def save_meal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    meal_name = update.message.text.strip()
    day = context.user_data['meal_day']
    meal_type = context.user_data.get('meal_type', 'meal')
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    c.execute(
        "INSERT INTO meals (name, type, day_of_week) VALUES (?, ?, ?)",
        (meal_name, meal_type, day)
    )
    conn.commit()
    conn.close()
    
    days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    meal_type_fa = 'غذا' if meal_type == 'meal' else 'دسر'
    
    await update.message.reply_text(
        f"✅ {meal_type_fa} '{meal_name}' برای روز {days[day]} اضافه شد!\n\n"
        "برای بازگشت به منو /start را ارسال کنید."
    )
    
    return ConversationHandler.END

# لیست غذاها
async def list_meals(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    text = "📋 لیست غذاها و دسرها:\n\n"
    
    for i, day in enumerate(days):
        text += f"📅 {day}:\n"
        
        c.execute("SELECT name FROM meals WHERE day_of_week = ? AND type = 'meal'", (i,))
        meals = c.fetchall()
        if meals:
            text += "  🍽 غذاها: " + ", ".join([m[0] for m in meals]) + "\n"
        
        c.execute("SELECT name FROM meals WHERE day_of_week = ? AND type = 'dessert'", (i,))
        desserts = c.fetchall()
        if desserts:
            text += "  🍰 دسرها: " + ", ".join([d[0] for d in desserts]) + "\n"
        
        text += "\n"
    
    conn.close()
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data='admin_meals')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)

# رزرو غذا توسط کاربران
async def reserve_food_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    # نمایش 14 روز آینده
    keyboard = []
    today = datetime.now().date()
    
    for i in range(14):
        date = today + timedelta(days=i)
        day_name = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه'][date.weekday()]
        date_str = date.strftime('%Y-%m-%d')
        button_text = f"{day_name} - {date.strftime('%d/%m')}"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=f'reserve_{date_str}')])
    
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='back_to_main')])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "روز مورد نظر برای رزرو را انتخاب کنید:",
        reply_markup=reply_markup
    )

# انتخاب غذا برای رزرو
async def select_meal_for_reservation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    date_str = query.data.split('_')[1]
    date = datetime.strptime(date_str, '%Y-%m-%d').date()
    day_of_week = date.weekday()
    
    context.user_data['reservation_date'] = date_str
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    c.execute("SELECT id, name FROM meals WHERE day_of_week = ? AND type = 'meal'", (day_of_week,))
    meals = c.fetchall()
    conn.close()
    
    if not meals:
        await query.edit_message_text(
            "❌ برای این روز غذایی تعریف نشده است.\n\n"
            "برای بازگشت /start را ارسال کنید."
        )
        return
    
    keyboard = []
    for meal_id, meal_name in meals:
        keyboard.append([InlineKeyboardButton(meal_name, callback_data=f'meal_{meal_id}')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='reserve_food')])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    day_name = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه'][day_of_week]
    await query.edit_message_text(
        f"غذای خود را برای {day_name} انتخاب کنید:",
        reply_markup=reply_markup
    )

# انتخاب دسر برای رزرو
async def select_dessert_for_reservation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    meal_id = int(query.data.split('_')[1])
    context.user_data['selected_meal_id'] = meal_id
    
    date_str = context.user_data['reservation_date']
    date = datetime.strptime(date_str, '%Y-%m-%d').date()
    day_of_week = date.weekday()
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    c.execute("SELECT id, name FROM meals WHERE day_of_week = ? AND type = 'dessert'", (day_of_week,))
    desserts = c.fetchall()
    conn.close()
    
    keyboard = []
    for dessert_id, dessert_name in desserts:
        keyboard.append([InlineKeyboardButton(dessert_name, callback_data=f'dessert_{dessert_id}')])
    keyboard.append([InlineKeyboardButton("بدون دسر", callback_data='dessert_none')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data=f'reserve_{date_str}')])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "دسر خود را انتخاب کنید:",
        reply_markup=reply_markup
    )

# تکمیل رزرو
async def complete_reservation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    dessert_id = None if query.data == 'dessert_none' else int(query.data.split('_')[1])
    meal_id = context.user_data['selected_meal_id']
    date_str = context.user_data['reservation_date']
    user_id = update.effective_user.id
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    try:
        c.execute(
            "INSERT OR REPLACE INTO reservations (user_id, meal_id, dessert_id, reservation_date) VALUES (?, ?, ?, ?)",
            (user_id, meal_id, dessert_id, date_str)
        )
        conn.commit()
        
        # دریافت نام غذا و دسر
        c.execute("SELECT name FROM meals WHERE id = ?", (meal_id,))
        meal_name = c.fetchone()[0]
        
        dessert_name = "بدون دسر"
        if dessert_id:
            c.execute("SELECT name FROM meals WHERE id = ?", (dessert_id,))
            dessert_name = c.fetchone()[0]
        
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        day_name = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه'][date.weekday()]
        
        await query.edit_message_text(
            f"✅ رزرو شما ثبت شد!\n\n"
            f"📅 روز: {day_name} - {date.strftime('%d/%m/%Y')}\n"
            f"🍽 غذا: {meal_name}\n"
            f"🍰 دسر: {dessert_name}\n\n"
            "برای بازگشت به منو /start را ارسال کنید."
        )
    except Exception as e:
        await query.edit_message_text(
            f"❌ خطا در ثبت رزرو: {str(e)}\n\n"
            "برای بازگشت /start را ارسال کنید."
        )
    finally:
        conn.close()

# مشاهده رزروهای کاربر
async def my_reservations(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    c.execute('''
        SELECT r.reservation_date, m1.name, m2.name
        FROM reservations r
        LEFT JOIN meals m1 ON r.meal_id = m1.id
        LEFT JOIN meals m2 ON r.dessert_id = m2.id
        WHERE r.user_id = ? AND r.reservation_date >= date('now')
        ORDER BY r.reservation_date
    ''', (user_id,))
    
    reservations = c.fetchall()
    conn.close()
    
    if not reservations:
        text = "شما هیچ رزروی ندارید."
    else:
        text = "📋 رزروهای شما:\n\n"
        for date_str, meal_name, dessert_name in reservations:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            day_name = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه'][date.weekday()]
            dessert_text = dessert_name if dessert_name else "بدون دسر"
            text += f"📅 {day_name} {date.strftime('%d/%m')}\n"
            text += f"   🍽 {meal_name}\n"
            text += f"   🍰 {dessert_text}\n\n"
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data='back_to_main')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)

# مشاهده رزروها توسط ادمین
async def admin_view_reservations(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    c.execute('''
        SELECT u.first_name, u.last_name, r.reservation_date, m1.name, m2.name
        FROM reservations r
        JOIN users u ON r.user_id = u.user_id
        LEFT JOIN meals m1 ON r.meal_id = m1.id
        LEFT JOIN meals m2 ON r.dessert_id = m2.id
        WHERE r.reservation_date >= date('now')
        ORDER BY r.reservation_date, u.first_name
    ''')
    
    reservations = c.fetchall()
    conn.close()
    
    if not reservations:
        text = "هیچ رزروی ثبت نشده است."
    else:
        text = "📊 رزروهای ثبت شده:\n\n"
        current_date = None
        for first_name, last_name, date_str, meal_name, dessert_name in reservations:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            if date != current_date:
                day_name = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه'][date.weekday()]
                text += f"\n📅 {day_name} {date.strftime('%d/%m/%Y')}:\n"
                current_date = date
            
            dessert_text = dessert_name if dessert_name else "بدون دسر"
            text += f"• {first_name} {last_name}: {meal_name} + {dessert_text}\n"
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data='back_to_admin')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)

# خروجی اکسل
async def export_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("در حال تولید فایل اکسل...")
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    
    # ایجاد Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "برنامه غذایی"
    
    # تنظیمات استایل
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # سرستون‌ها
    today = datetime.now().date()
    headers = ["نام"]
    dates = []
    
    for i in range(14):
        date = today + timedelta(days=i)
        day_name = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه'][date.weekday()]
        headers.append(f"{day_name}\n{date.strftime('%d/%m')}")
        dates.append(date.strftime('%Y-%m-%d'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # دریافت کاربران
    c.execute("SELECT user_id, first_name, last_name FROM users WHERE is_active = 1 ORDER BY first_name")
    users = c.fetchall()
    
    # پر کردن داده‌ها
    for row, (user_id, first_name, last_name) in enumerate(users, 2):
        ws.cell(row=row, column=1, value=f"{first_name} {last_name}").alignment = center_alignment
        
        for col, date_str in enumerate(dates, 2):
            c.execute('''
                SELECT m1.name, m2.name
                FROM reservations r
                LEFT JOIN meals m1 ON r.meal_id = m1.id
                LEFT JOIN meals m2 ON r.dessert_id = m2.id
                WHERE r.user_id = ? AND r.reservation_date = ?
            ''', (user_id, date_str))
            
            result = c.fetchone()
            if result:
                meal_name, dessert_name = result
                cell_value = meal_name
                if dessert_name:
                    cell_value += f"\n{dessert_name}"
                ws.cell(row=row, column=col, value=cell_value).alignment = center_alignment
    
    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 20
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    
    # ذخیره فایل
    filename = f"food_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    conn.close()
    
    # ارسال فایل
    with open(filename, 'rb') as file:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=file,
            filename=filename,
            caption="📊 برنامه غذایی دو هفته آینده"
        )
    
    os.remove(filename)
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data='back_to_admin')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.reply_text("فایل اکسل ارسال شد.", reply_markup=reply_markup)

# ارسال پیام همگانی
async def start_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "📢 پیام خود را برای ارسال به همه کاربران وارد کنید:\n\n"
        "برای لغو /cancel را ارسال کنید."
    )
    
    return BROADCAST_MSG

# ارسال پیام به همه
async def send_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message.text
    
    conn = sqlite3.connect('food_reservation.db')
    c = conn.cursor()
    c.execute("SELECT user_id FROM users WHERE is_active = 1")
    users = c.fetchall()
    conn.close()
    
    success_count = 0
    fail_count = 0
    
    for (user_id,) in users:
        try:
            await context.bot.send_message(chat_id=user_id, text=f"📢 پیام از مدیریت:\n\n{message}")
            success_count += 1
        except Exception as e:
            logger.error(f"Failed to send to {user_id}: {e}")
            fail_count += 1
    
    await update.message.reply_text(
        f"✅ پیام شما به {success_count} کاربر ارسال شد.\n"
        f"❌ {fail_count} کاربر دریافت نکردند.\n\n"
        "برای بازگشت به منو /start را ارسال کنید."
    )
    
    return ConversationHandler.END

# لغو عملیات
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "عملیات لغو شد.\n\n"
        "برای بازگشت به منو /start را ارسال کنید."
    )
    return ConversationHandler.END

# هندلر callback query
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    
    if not is_admin(update.effective_user.id) and not is_authorized_user(update.effective_user.id):
        await query.answer("شما دسترسی ندارید.", show_alert=True)
        return
    
    data = query.data
    
    # مسیریابی
    if data == 'back_to_admin':
        await admin_menu(update, context)
    elif data == 'back_to_main':
        await query.answer()
        keyboard = [
            [InlineKeyboardButton("🍽 رزرو غذا", callback_data='reserve_food')],
            [InlineKeyboardButton("📋 رزروهای من", callback_data='my_reservations')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("منوی اصلی:", reply_markup=reply_markup)
    elif data == 'admin_users':
        await admin_users_menu(update, context)
    elif data == 'admin_meals':
        await admin_meals_menu(update, context)
    elif data == 'list_users':
        await list_users(update, context)
    elif data == 'list_meals':
        await list_meals(update, context)
    elif data == 'add_meal':
        await select_day_for_meal(update, context)
    elif data == 'add_dessert':
        await select_day_for_dessert(update, context)
    elif data.startswith('day_meal_'):
        await receive_meal_day(update, context)
    elif data.startswith('day_dessert_'):
        await receive_meal_day(update, context)
    elif data == 'admin_view_reservations':
        await admin_view_reservations(update, context)
    elif data == 'admin_export_excel':
        await export_to_excel(update, context)
    elif data == 'reserve_food':
        await reserve_food_menu(update, context)
    elif data.startswith('reserve_'):
        await select_meal_for_reservation(update, context)
    elif data.startswith('meal_'):
        await select_dessert_for_reservation(update, context)
    elif data.startswith('dessert_'):
        await complete_reservation(update, context)
    elif data == 'my_reservations':
        await my_reservations(update, context)

def main():
    # دریافت توکن از متغیر محیطی
    TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
    PORT = int(os.getenv('PORT', 8443))
    
    if not TOKEN:
        logger.error("TELEGRAM_BOT_TOKEN not found!")
        return
    
    # ایجاد دیتابیس
    init_db()
    
    # ایجاد Application
    application = Application.builder().token(TOKEN).build()
    
    # ConversationHandler برای افزودن کاربر
    add_user_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(start_add_user, pattern='^add_user$')],
        states={
            ADD_USER_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_user_id)],
            ADD_USER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_user_name)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    # ConversationHandler برای افزودن غذا
    add_meal_handler = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(receive_meal_day, pattern='^day_meal_'),
            CallbackQueryHandler(receive_meal_day, pattern='^day_dessert_')
        ],
        states={
            ADD_MEAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_meal)],
            ADD_DESSERT: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_meal)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    # ConversationHandler برای پیام همگانی
    broadcast_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(start_broadcast, pattern='^admin_broadcast$')],
        states={
            BROADCAST_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, send_broadcast)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    # اضافه کردن handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(add_user_handler)
    application.add_handler(add_meal_handler)
    application.add_handler(broadcast_handler)
    application.add_handler(CallbackQueryHandler(button_handler))
    
    # راه‌اندازی
    if os.getenv('RAILWAY_ENVIRONMENT'):
        # حالت webhook برای Railway
        WEBHOOK_URL = os.getenv('RAILWAY_PUBLIC_DOMAIN')
        application.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            url_path=TOKEN,
            webhook_url=f"https://{WEBHOOK_URL}/{TOKEN}"
        )
    else:
        # حالت polling برای تست محلی
        application.run_polling()

if __name__ == '__main__':
    main()
